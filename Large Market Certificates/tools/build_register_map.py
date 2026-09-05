"""Add the Envizi account mapping to the site register, rebuilt from the pristine source."""
import csv, json, shutil, warnings, zipfile, os, re, datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import FormulaRule
warnings.filterwarnings("ignore")

OUTDIR = "/home/user/RevenueAndUnallocated/Large Market Certificates"
SRC = f"{OUTDIR}/Downer_Energy_Contracting_and_Budget_Summary_FY26-28.xlsx"
DST = f"{OUTDIR}/Downer_Energy_Contracting_and_Budget_Summary_FY26-28_with_Envizi_accounts.xlsx"
ACC = "/home/user/RevenueAndUnallocated/FY27/Extract_for_Accounts 05 Sep 26.csv"
LOC = "/home/user/RevenueAndUnallocated/FY27/Extract_for_Locations 26 Aug 26.csv"
GUIDE = os.path.join(os.path.dirname(__file__), "guide_data.json")
ZERO = "30 Dec 1899"

HEADERS = ["Envizi Account Number", "Envizi Account Style", "Envizi Data Type", "Envizi Supplier",
           "Envizi Location", "Envizi Location Ref", "Account Status", "Match Basis",
           "Other Envizi accounts on this ID", "Certificate / virtual meter account"]
WIDTHS = [34, 26, 24, 16, 34, 16, 20, 34, 46, 34]
FONT = "Aptos Narrow"
F_HEAD = Font(name=FONT, size=11, bold=True, color="FFFFFFFF")
F_BASE = Font(name=FONT, size=11)
FILL_HEAD = PatternFill("solid", start_color="FF1F3864", end_color="FF1F3864")
THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def nmi_of(a):
    a = a[:-len("_CERTS")] if a.endswith("_CERTS") else a      # the NZ Eco_ICP_<icp>_CERTS pattern
    return a.rsplit("_", 1)[-1] if "_" in a else a

rows = list(csv.reader(open(ACC, newline="", encoding="utf-8-sig")))
H, body = rows[0], rows[1:]
col = {n: i for i, n in enumerate(H)}
accounts = []
for rec in body:
    d = {n: rec[i] for n, i in col.items()}
    d["_nmi"] = nmi_of(d["Account Number"])
    d["_rep"] = "" if d["Replaced On"].strip() in ("", ZERO) else d["Replaced On"].strip()
    try:
        future = d["_rep"] != "" and datetime.datetime.strptime(d["_rep"], "%d %b %Y") > datetime.datetime(2026, 9, 5)
    except ValueError:
        future = False
    d["_active"] = (d["_rep"] == "" or future) and d["Location"] != "Unallocated Accounts"
    accounts.append(d)

loc_ref = {}
for rec in list(csv.reader(open(LOC, newline="", encoding="utf-8-sig")))[1:]:
    loc_ref.setdefault(rec[1], rec[4])

guide = json.load(open(GUIDE))
new_by_nmi = {r["nmi"]: r["new_acct"] for r in guide["rows"] if r["decision"] == "Create"}
src_by_nmi = {r["nmi"]: r["src_acct"] for r in guide["rows"] if r["decision"] == "Create"}
live = {l["account"] for l in guide["live"]}
dual_nmis = {d["nmi"] for d in guide["dual"]}
waiting_by_nmi = {w["nmi"]: w for w in guide.get("waiting", [])}

by_nmi = {}
for a in accounts:
    by_nmi.setdefault(a["_nmi"], []).append(a)
NEW_ACCTS = set(new_by_nmi.values())
certs_by_loc = {}
for a in accounts:
    if (a["Data Type"] == "Certificates - Location [kWh]" and a["_active"]
            and a["Account Number"] not in NEW_ACCTS):
        certs_by_loc.setdefault(a["Location"], []).append(a)

shutil.copy(SRC, DST)
wb = openpyxl.load_workbook(DST)
S = wb["Site Register"]
LAST = S.max_row

for j, (h, w) in enumerate(zip(HEADERS, WIDTHS)):
    c = S.cell(2, 22 + j, h)
    c.font, c.fill, c.border = F_HEAD, FILL_HEAD, BORDER
    c.alignment = Alignment(wrap_text=True, vertical="center")
    S.column_dimensions[openpyxl.utils.get_column_letter(22 + j)].width = w

matched = 0
for r in range(3, LAST + 1):
    cid = str(S.cell(r, 1).value or "").strip()
    if not cid:
        continue
    cands = by_nmi.get(cid, [])
    # the account the register row maps to: the one this NMI's virtual meter follows where there is
    # one, else the active account for the row's own commodity, else whatever is on the ID
    commodity = str(S.cell(r, 2).value or "").strip()
    want = {"Electricity": "Electricity", "Natural Gas": "Natural Gas"}.get(commodity, "")
    def rank(a):
        return (0 if a["Account Number"] == src_by_nmi.get(cid) else 1,
                0 if (want and a["Data Type"].startswith(want)) else 1,
                0 if a["_active"] else 1,
                a["Account Number"])
    pick = sorted(cands, key=rank)[0] if cands else None
    if pick:
        matched += 1
        others = "; ".join(a["Account Number"] for a in cands if a is not pick)
        vals = [pick["Account Number"], pick["Account Style"], pick["Data Type"], pick["Supplier"],
                pick["Location"], loc_ref.get(pick["Location"], ""),
                "Active" if pick["_active"] else f"Replaced {pick['_rep']}",
                "Connection ID is the account number suffix", others]
    else:
        vals = ["Not found", "", "", "", "", "", "", "No Envizi account ends in this Connection ID", ""]

    cert = []
    if cid in new_by_nmi:
        n = new_by_nmi[cid]
        cert.append(f"{n} (live)" if n in live else f"{n} (to create)")
    if cid in waiting_by_nmi:
        w = waiting_by_nmi[cid]
        cert.append(f"{w['new_acct']} (TEMPORARY - remake against the {w['expected_supplier']} account when it "
                    + ("appears; " + w["parked"] + " is sitting at Unallocated Accounts, allocate it first)"
                       if w["parked"] else "appears; not created yet)"))
    if pick:
        for c in certs_by_loc.get(pick["Location"], []):
            tag = "" if c["_nmi"] == cid else " (at this location, different ID)"
            cert.append(f"{c['Account Number']}{tag} - historical")
    vals.append(" · ".join(cert))

    for k, v in enumerate(vals):
        c = S.cell(r, 22 + k, v)
        c.font, c.border = F_BASE, BORDER
        c.alignment = Alignment(vertical="top", wrap_text=(k in (8, 9)))

RNG = f"V3:AE{LAST}"
S.conditional_formatting.add(RNG, FormulaRule(formula=['$V3="Not found"'],
    fill=PatternFill("solid", start_color="FFF8CBAD", end_color="FFF8CBAD"), stopIfTrue=False))
S.conditional_formatting.add(f"AB3:AB{LAST}", FormulaRule(formula=['AND($AB3<>"Active",$AB3<>"")'],
    fill=PatternFill("solid", start_color="FFFFF2CC", end_color="FFFFF2CC"), stopIfTrue=False))
S.conditional_formatting.add(f"AD3:AD{LAST}", FormulaRule(formula=[f'AND($AD3<>"",COUNTIF(DoubleCount,$A3)>0)'],
    fill=PatternFill("solid", start_color="FFFCE4D6", end_color="FFFCE4D6"), stopIfTrue=False))
S.conditional_formatting.add(f"AE3:AE{LAST}", FormulaRule(formula=['ISNUMBER(SEARCH("(live)",$AE3))'],
    fill=PatternFill("solid", start_color="FFE2EFDA", end_color="FFE2EFDA"), stopIfTrue=False))
S.conditional_formatting.add(f"AE3:AE{LAST}", FormulaRule(formula=['ISNUMBER(SEARCH("TEMPORARY",$AE3))'],
    fill=PatternFill("solid", start_color="FFFFF2CC", end_color="FFFFF2CC"), stopIfTrue=False))
S.auto_filter.ref = f"A2:AE{LAST}"
S.freeze_panes = "B3"

# a named range so the double-count rule above is readable, parked out of the way
ws = wb.create_sheet("Envizi mapping notes")
ws["A1"] = "How columns V to AE were filled"
ws["A1"].font = Font(name=FONT, size=14, bold=True)
notes = [
    "Connection ID is matched to the text after the last underscore in an Envizi account number.",
    "Where the site has a certificate virtual meter, columns V to AC describe the account that meter",
    "follows - the one carrying the electricity from 1 July 2026. Otherwise they describe the active",
    "account on that ID. Column AD lists every other Envizi account on the same ID.",
    "",
    "Column AE names the certificate account: the new virtual meter (live, or still to create) and any",
    "historical LGCS_ account at the same location. Green means it is already live in Envizi; amber means the",
    "certificate account is TEMPORARY - the contracted retailer has no account on that ID yet, so the meter",
    "follows the old supply and has to be deleted and remade once the right account appears.",
    "",
    "Shading: red in V to AE means no Envizi account ends in that Connection ID. Amber in AB means the",
    "account is closed. Peach in AD flags a Connection ID that is being double counted - a second",
    "account is still recording the months the source account already covers (list below).",
    "",
    "Scope note: large and small market is a metering classification and does not decide whether the",
    "renewal covers a site - Category Management confirmed on 4 September 2026 that a site can be",
    "metered small market and still be supplied as large market. The site register's green rows decide.",
    "",
    "Built from Extract_for_Accounts 05 Sep 26.csv and Extract_for_Locations 26 Aug 26.csv.",
]
for i, t in enumerate(notes, start=3):
    ws.cell(i, 1, t).font = Font(name=FONT, size=11)
ws.column_dimensions["A"].width = 110
ws["A22"] = "Connection IDs being double counted"
ws["A22"].font = Font(name=FONT, size=11, bold=True)
for i, n in enumerate(sorted(dual_nmis), start=23):
    ws.cell(i, 1, n).font = Font(name=FONT, size=11)
wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
    "DoubleCount", attr_text=f"'Envizi mapping notes'!$A$23:$A${22 + len(dual_nmis)}"))

wb.properties.creator = ""
wb.properties.lastModifiedBy = ""
wb.properties.title = ""
wb.save(DST)

# strip the generator's name out of the app metadata as well
tmp = DST + ".tmp"
with zipfile.ZipFile(DST) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
    for it in zin.infolist():
        d = zin.read(it.filename)
        if it.filename == "docProps/app.xml":
            d = re.sub(rb"<Application>.*?</Application>", b"<Application></Application>", d)
        zout.writestr(it, d)
os.replace(tmp, DST)
print(f"matched {matched} of 341 register rows | certificate accounts named: "
      f"{sum(1 for r in range(3, LAST+1) if '_CERTS' in str(S.cell(r,31).value or ''))}")
