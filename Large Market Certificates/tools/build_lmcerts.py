"""Large market renewable-certificate virtual accounts - Envizi Account Setup and Data Load (PM&C) build."""
import csv, datetime, os
from collections import OrderedDict

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as L

ROOT   = "/home/user/RevenueAndUnallocated"
OUTDIR = os.path.join(ROOT, "Large Market Certificates")
OUT    = os.path.join(OUTDIR, "Account_Setup_and_Data_Load_-_PM&C_LMCERTSJUL26_Setup.xlsx")
SRC_REG  = os.path.join(OUTDIR, "Downer_Energy_Contracting_and_Budget_Summary_FY26-28.xlsx")
SRC_SUMM = os.path.join(OUTDIR, "ElectricityEnviziSummaryjunejulyaug26.xlsx")
SRC_ACC  = os.path.join(ROOT, "FY27/Extract_for_Accounts 05 Sep 26.csv")
SRC_LOC  = os.path.join(ROOT, "FY27/Extract_for_Locations 26 Aug 26.csv")

S_LOAD, S_PREP, S_REV, S_CHK, S_NOTES = ("Account_Setup_and_Data_Load", "Prep - with formulas",
                                         "Review - AU Large Market", "Check", "Notes")
S_MAN, S_LGC = "Manual Setup Checklist", "LGCS Accounts to Check"
S_REG, S_ACC, S_LOC, S_SUM = ("Site Register (AU Large)", "Extract for Accounts 05 Sep 26",
                              "Extract_for_Locations 26 Aug 26", "Envizi Summary Jun-Aug 26")
q = lambda s: "'" + s + "'"
REV = q(S_REV)

ORG_LINK, ORG = 37395, "Downer"
STYLE_CAPTION = "Certificates - Location - kWh"
CERT_TYPE     = "Certificates - Location [kWh]"
GREEN_TYPE    = "Electricity - Green [kWh]"
ELEC_TYPE     = "Electricity [kWh]"
ZERO_DATE     = "30 Dec 1899"
SUFFIX        = "_CERTS"
PAT_ECO, PAT_LGC = "Ecotricity - account_CERTS", "AU LGC - LGCS_NMI"
LGC_HIST, LGC_EMPTY = "Historical - hold data, create a new account", "Empty - reuse as the virtual meter (exclude here)"
NEW_SUPPLIER = "LGC Virtual Account"
RENEWAL_START = datetime.datetime(2026, 7, 1)
REC_REF       = "Renewable certificates - virtual account setup"
SUPPLIER_MAP  = OrderedDict([("Engie", "EngieAU"), ("Origin", "Origin"), ("Shell", "ShellEnergyAU"),
                             ("Alinta", "Alinta"), ("CS Energy", "CSEnergy"), ("Aurora", "Aurora"),
                             ("Jacana Energy", "Jacana"), ("Jacana Energy (Standing)", "Jacana")])
NAMED_EXCL    = OrderedDict([("3053253239", "QTMP (Torbanlea)"), ("2500033707", "Northern Territory"),
                             ("2500044629", "Northern Territory"), ("2500054287", "Northern Territory")])
D_CREATE, D_LGC, D_GREEN, D_NAMED, D_HOLD = ("Create", "Exclude - offset exists (LGC account)",
    "Exclude - offset on account green component", "Exclude - named site (NT / QTMP / Pakenham)", "Hold")
LM_STYLES = ("Electricity Large Market", "Energetics - Large Market")
D_TEMP = "Create - temporary, retailer's account not in Envizi yet"
MAKE = (D_CREATE, D_TEMP)          # both get an account built; D_TEMP gets remade later
DECISIONS = [D_CREATE, D_TEMP, D_LGC, D_GREEN, D_NAMED, D_HOLD]

# ------------------------------------------------------------------ styling
FONT = "Aptos Narrow"
F_BASE  = Font(name=FONT, size=11)
F_BOLD  = Font(name=FONT, size=11, bold=True)
F_TITLE = Font(name=FONT, size=14, bold=True)
F_HEAD  = Font(name=FONT, size=11, bold=True, color="FFFFFFFF")
F_NOTE  = Font(name=FONT, size=10, italic=True, color="FF595959")
F_LINK  = Font(name=FONT, size=11, color="FF005A87")
FILL_HEAD  = PatternFill("solid", start_color="FF1F3864", end_color="FF1F3864")
FILL_SUB   = PatternFill("solid", start_color="FFD9E1F2", end_color="FFD9E1F2")
FILL_INPUT = PatternFill("solid", start_color="FFFFFACD", end_color="FFFFFACD")
FILL_GREEN = PatternFill("solid", start_color="FFE2EFDA", end_color="FFE2EFDA")
FILL_GREY  = PatternFill("solid", start_color="FFEDEDED", end_color="FFEDEDED")
FILL_AMBER = PatternFill("solid", start_color="FFFFF2CC", end_color="FFFFF2CC")
FILL_RED   = PatternFill("solid", start_color="FFF8CBAD", end_color="FFF8CBAD")
THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP = Alignment(vertical="top")


def hdr(ws, row, values, fill=FILL_HEAD, font=F_HEAD, height=30):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row, c, v)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, CENTER, BORDER
    ws.row_dimensions[row].height = height


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def parse_date(raw):
    raw = (raw or "").strip()
    if not raw or raw == ZERO_DATE:
        return None
    for fmt in ("%d %b %Y", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(raw, fmt)
        except ValueError:
            pass
    return None


def nmi_of(acct):
    return acct.rsplit("_", 1)[-1] if "_" in acct else acct


# ------------------------------------------------------------------ sources
reg_wb  = openpyxl.load_workbook(SRC_REG)
reg_wbv = openpyxl.load_workbook(SRC_REG, data_only=True)
ws_r, ws_rv = reg_wb["Site Register"], reg_wbv["Site Register"]
REG_HEADERS = [ws_r.cell(2, c).value for c in range(1, 22)]

def is_green(r):
    c = ws_r.cell(r, 1)
    return bool(c.fill and c.fill.fill_type == "solid" and c.fill.start_color.type == "rgb"
                and c.fill.start_color.rgb == "FF92D050")

register = []
for r in range(3, ws_r.max_row + 1):
    if ws_r.cell(r, 5).value != "AU Large Electricity":
        continue
    vals = [ws_rv.cell(r, c).value for c in range(1, 22)]
    vals[0] = str(vals[0]).strip()
    register.append(dict(src_row=r, green=is_green(r), vals=vals))
print("register AU Large Electricity rows:", len(register), "green:", sum(x["green"] for x in register))
REG_NMIS = [x["vals"][0] for x in register]

with open(SRC_ACC, newline="", encoding="utf-8-sig") as fh:
    acc_rows = list(csv.reader(fh))
ACC_HEADERS, acc_body = acc_rows[0], acc_rows[1:]
acc = pd.DataFrame(acc_body, columns=ACC_HEADERS)
acc["_nmi"] = acc["Account Number"].map(nmi_of)
acc["_rep"] = acc["Replaced On"].str.strip().replace(ZERO_DATE, "")
EXTRACT_DAY = datetime.datetime(2026, 9, 5)
def _still_open(rep):
    if rep == "":
        return True
    try:
        return datetime.datetime.strptime(rep, "%d %b %Y") > EXTRACT_DAY      # future close date = open today
    except ValueError:
        return False
acc["_active"] = (acc["_rep"].map(_still_open) & (acc["Location"] != "Unallocated Accounts")).astype(int)
keep = acc[((acc["Data Type"] == ELEC_TYPE) & acc["_nmi"].isin(REG_NMIS)) | (acc["Data Type"] == CERT_TYPE)].copy()
keep["_pref"] = keep["Account Style"].str.contains("Large Market").map({True: 0, False: 1})
keep["_kind"]  = (keep["Data Type"] == CERT_TYPE).astype(int)

with open(SRC_LOC, newline="", encoding="utf-8-sig") as fh:
    loc_rows = list(csv.reader(fh))
LOC_HEADERS, loc_body = loc_rows[0], loc_rows[1:]
loc_ref = {}
for rec in loc_body:
    loc_ref.setdefault(rec[1], rec[4])

summ = pd.read_excel(SRC_SUMM).fillna("")
summ["_nmi"] = summ["Item Number"].astype(str).map(nmi_of)
summ_keep = summ[summ["_nmi"].isin(REG_NMIS)].drop(columns=["_nmi"])
print("summary tab rows:", len(summ_keep))

# Which account actually carries each NMI's electricity from 1 Jul 26. The renewal follows the supply
# agreement, not the meter class - Nick Goldsworthy confirmed on 4 Sep 26 that a site can be metered
# small market and still be supplied as large market - so the register's green row decides scope and
# the Envizi account style only breaks a tie. Ordering the accounts tab this way means the first
# active electricity row for an NMI is the source, in the workbook and here alike.
_post = summ[(summ["Data Type"] == ELEC_TYPE)
             & (pd.to_datetime(summ["Occurred_On"], errors="coerce") >= RENEWAL_START)]
_post = _post.groupby(_post["Item Number"].astype(str))["Actual Data"].apply(
    lambda c: float(pd.to_numeric(c, errors="coerce").fillna(0).sum()))
keep["_post"] = (keep["Account Number"].map(_post).fillna(0) > 0).astype(int)
keep = keep.sort_values(["_kind", "_nmi", "_active", "_post", "_pref", "Account Number"],
                        ascending=[True, True, False, False, True, True])
print("accounts tab rows:", len(keep), " (elec", (keep["_kind"] == 0).sum(), "/ certs", (keep["_kind"] == 1).sum(), ")")
ACC_N = 1 + len(keep); LOC_N = 1 + len(loc_body); SUM_N = 1 + len(summ_keep)

# ------------------------------------------------------------------ the match (python side, for values + checking)
elec_act = keep[(keep["_kind"] == 0) & (keep._active == 1)]
certs_act = keep[(keep["_kind"] == 1) & (keep._active == 1) & (keep["Supplier"] != NEW_SUPPLIER)]
review = []
for i, x in enumerate(register):
    v = x["vals"]; nmi = v[0]
    m = elec_act[elec_act._nmi == nmi]
    a = m.iloc[0] if len(m) else None
    location = a["Location"] if a is not None else "Not found"
    lref = loc_ref.get(location, "Not in locations extract") if a is not None else ""
    c_loc = certs_act[certs_act["Location"] == location] if a is not None else certs_act.iloc[0:0]
    c_nmi = c_loc[c_loc._nmi == nmi]
    s = summ_keep[summ_keep["Item Number"].astype(str) == (a["Account Number"] if a is not None else "~")]
    s_ren = s[pd.to_datetime(s["Occurred_On"], errors="coerce") >= RENEWAL_START]
    green_kwh = float(pd.to_numeric(s_ren[s_ren["Data Type"] == GREEN_TYPE]["Total Data"], errors="coerce").fillna(0).sum())
    green_co2 = float(pd.to_numeric(s_ren[s_ren["Data Type"] == GREEN_TYPE]["Total CO2e(t)"], errors="coerce").fillna(0).sum())
    cons_kwh  = float(pd.to_numeric(s[s["Data Type"] == ELEC_TYPE]["Total Data"], errors="coerce").fillna(0).sum())
    contracted = v[12]; start = v[13]
    env_sup = a["Supplier"] if a is not None else ""
    new_sup = NEW_SUPPLIER
    ret_sup = SUPPLIER_MAP.get(contracted, contracted)
    src = a                              # keep is pre-ranked, so the matched account is the source
    src_acct = src["Account Number"] if src is not None else "None"
    src_small = bool(src is not None and src["Account Style"] not in LM_STYLES)
    other_on_nmi = [f"{r['Account Number']} ({r['Supplier']}, {r['Account Style']})"
                    for _, r in m.iterrows() if r["Account Number"] != src_acct]
    s_src = summ_keep[summ_keep["Item Number"].astype(str) == src_acct]
    pre = s_src[(s_src["Data Type"] == ELEC_TYPE) & (pd.to_datetime(s_src["Occurred_On"], errors="coerce") < RENEWAL_START)]
    src_history = "Yes" if float(pd.to_numeric(pre["Total Data"], errors="coerce").fillna(0).sum()) > 0 else "No"
    named = NAMED_EXCL.get(nmi, "")
    if named:
        decision = D_NAMED
    elif src is None:
        decision = D_HOLD
    elif green_kwh > 0:
        decision = D_GREEN
    elif contracted and contracted not in str(env_sup) and not any(
            SUPPLIER_MAP.get(contracted, contracted) == r["Supplier"] for _, r in m.iterrows()):
        # contracted to a retailer that has no account on this NMI yet - the connector has not
        # switched the site, so a virtual meter built now would mirror the wrong supply
        decision = D_TEMP
    elif lref in ("", "Not in locations extract"):
        decision = D_HOLD
    else:
        decision = D_CREATE
    # reason text
    if decision == D_NAMED:
        reason = f"Named exclusion - {named}."
        if named.startswith("QTMP"):
            reason += " The account sits at 'Torbanlea - QTMP', which is also absent from the 26 Aug 26 locations extract."
        else:
            reason += " Jacana Energy site; not green in the Site Register either."
    elif decision == D_LGC:
        names = ", ".join(c_loc["Account Number"])
        if len(c_nmi):
            reason = f"LGC account already at this location on this NMI: {names}."
        else:
            reason = (f"{names} sits at this location but on a different NMI to the register's {nmi} - excluded on "
                      "the location rule. Confirm that LGC account still relates to this site before revisiting.")
    elif decision == D_GREEN:
        reason = (f"{contracted} contract from {start:%d %b %Y} already records green kWh on this account's own green "
                  f"component (Jul-Aug 26: {green_kwh:,.0f} kWh, {green_co2:,.1f} tCO2e) - an offset is in place "
                  "without a separate certificate account. Flip to Create if a certificate account is wanted as well.")
    elif decision == D_TEMP:
        reason = (f"Contracted to {contracted} from {start:%d %b %Y}, but no {SUPPLIER_MAP.get(contracted, contracted)} "
                  f"account exists on this NMI in Envizi yet - the site is still recording on {env_sup} ({src_acct}). "
                  "Build the virtual account against that account now, but treat it as temporary: the name keys off the "
                  f"source, so when the {SUPPLIER_MAP.get(contracted, contracted)} account appears this one has to be "
                  "deleted and remade against it.")
    elif decision == D_HOLD:
        reason = ("No active electricity account on this NMI in Envizi - check before loading." if src is None
                  else "No location reference resolved - check before loading.")
    else:
        if len(c_loc):
            names = ", ".join(c_loc["Account Number"])
            reason = (f"Existing LGC account(s) here - {names} - hold 2025-and-earlier data (checked 4 Sep 26), so they cannot become a "
                      f"virtual meter and do not cover the renewal period. New virtual account needed. Contracted to {contracted} from {start:%d %b %Y}.")
        else:
            reason = f"No certificate account or green offset at this location. Contracted to {contracted} from {start:%d %b %Y}."
        if src_history == "Yes":
            reason += (f" Source account {src_acct} has data before 1 Jul 26 - date-bound the virtual meter to the renewal start or it will "
                       "generate certificates for periods that were not renewable.")
        if other_on_nmi:
            reason += (f" Note: {', '.join(other_on_nmi)} also records against this NMI - not the meter's source. "
                       "Close it off; see the double count list.")
        if nmi == "8000326927":
            reason += (" Shell's DEDI01_088_8000326927 is still at Unallocated Accounts (Sep-26 tracker row 5); "
                       "once allocated here it becomes the live account.")
    review.append(dict(idx=i + 1, src_row=x["src_row"], nmi=nmi, state=v[3], green="Y" if x["green"] else "N",
                       site=v[5], addr=v[6], bu=v[8], contracted=contracted, start=start,
                       account=a["Account Number"] if a is not None else "Not found", location=location, lref=lref,
                       style=a["Account Style"] if a is not None else "", env_sup=env_sup,
                       n_cert_loc=len(c_loc), n_cert_nmi=len(c_nmi), cert_names=", ".join(c_loc["Account Number"]),
                       green_kwh=green_kwh, cons_kwh=cons_kwh, decision=decision, reason=reason,
                       src_acct=src_acct, src_history=src_history, src_sup=(src["Supplier"] if src is not None else ""),
                       src_style=(src["Account Style"] if src is not None else ""), other_on_nmi="; ".join(other_on_nmi),
                       new_acct=(src_acct + SUFFIX) if decision in MAKE else "",
                       new_sup=new_sup if decision in MAKE else "",
                       rec_start=start if decision in MAKE else None,
                       rec_end=(start + pd.offsets.MonthEnd(0)).to_pydatetime() if decision in MAKE else None))
from collections import Counter
print("decisions:", Counter(r["decision"] for r in review))
creates = [r for r in review if r["decision"] in MAKE]
print("create rows:", len(creates))

# ================================================================== workbook
wb = openpyxl.Workbook()
ws_load = wb.active; ws_load.title = S_LOAD
ws_prep = wb.create_sheet(S_PREP)
ws_rev  = wb.create_sheet(S_REV)
ws_chk  = wb.create_sheet(S_CHK)
ws_man  = wb.create_sheet(S_MAN)
ws_lgc  = wb.create_sheet(S_LGC)
ws_not  = wb.create_sheet(S_NOTES)
ws_reg  = wb.create_sheet(S_REG)
ws_acc  = wb.create_sheet(S_ACC)
ws_loc  = wb.create_sheet(S_LOC)
ws_sum  = wb.create_sheet(S_SUM)

# ---------------------------------------------------------------- Review tab
R = ws_rev
R["A1"] = "AU large market sites - renewable certificate virtual accounts (built 05 Sep 2026)"
R["A1"].font = F_TITLE
R["A2"] = ("One row per AU Large Electricity connection in the Site Register. Formula columns read the extract tabs; "
           "yellow cells are inputs. Decision (col V) is pre-filled from Suggested (col U) and is yours to change - "
           "the row shades itself, and the Check tab flags any row where the two differ.")
R["A2"].font = F_NOTE; R.merge_cells("A2:N2"); R["A2"].alignment = WRAP; R.row_dimensions[2].height = 30

inputs = [("Organization Link", ORG_LINK), ("Organization", ORG),
          ("Account Style Link  (Certificates - Location - kWh)  - REQUIRED, from Envizi > Admin > Account Styles", None),
          ("Account Style Caption", STYLE_CAPTION), ("New account number - prefix", None),
          ("New account number - suffix", SUFFIX), ("Account Reader for the new accounts", None),
          ("Record Reference", REC_REF), ("Placeholder Quantity (kWh) on the setup record", 0),
          ("Record Entry Method", "Overwrite"),
          ("Account pattern  (Ecotricity: <account>_CERTS, supplier = contracted retailer  |  AU LGC: LGCS_<NMI>, supplier and reader LGCs - collides where an LGCS_ account already exists)", PAT_ECO),
          ("Existing LGC accounts at a location are...  (checked 4 Sep 26: all hold 2025-and-earlier data)", LGC_HIST),
          ("Supplier on the new accounts  (as used on the first two, 4 Sep 26)", NEW_SUPPLIER)]
R["A3"] = "Inputs"; R["A3"].font = F_BOLD
for i, (label, val) in enumerate(inputs):
    r = 4 + i
    R.cell(r, 1, label).font = F_BASE
    R.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = R.cell(r, 3, val); c.fill = FILL_INPUT; c.border = BORDER; c.font = F_BASE
IN_ORG, IN_ORGNAME, IN_STYLE, IN_CAPTION, IN_PREFIX, IN_SUFFIX, IN_READER, IN_REF, IN_QTY, IN_METHOD, IN_PATTERN, IN_LGCMODE, IN_NEWSUP = [f"$C${4+i}" for i in range(13)]
R["C6"].font = Font(name=FONT, size=11, bold=True, color="FFC00000")
dv_method = DataValidation(type="list", formula1='"Overwrite,Insert"', allow_blank=False)
R.add_data_validation(dv_method); dv_method.add("C13")
dv_pat = DataValidation(type="list", formula1=f'"{PAT_ECO},{PAT_LGC}"', allow_blank=False)
R.add_data_validation(dv_pat); dv_pat.add("C14")
dv_lgc = DataValidation(type="list", formula1=f'"{LGC_HIST},{LGC_EMPTY}"', allow_blank=False)
R.add_data_validation(dv_lgc); dv_lgc.add("C15")

R["E3"] = "Contracted retailer  ->  Envizi supplier"; R["E3"].font = F_BOLD
for i, (k, v) in enumerate(SUPPLIER_MAP.items()):
    R.cell(4 + i, 5, k).font = F_BASE
    c = R.cell(4 + i, 6, v); c.fill = FILL_INPUT; c.border = BORDER; c.font = F_BASE
MAP_RNG = (f"$E$4:$E${3+len(SUPPLIER_MAP)}", f"$F$4:$F${3+len(SUPPLIER_MAP)}")

R["H3"] = "Named exclusions (Connection ID  ->  reason)"; R["H3"].font = F_BOLD
for i, (k, v) in enumerate(NAMED_EXCL.items()):
    c = R.cell(4 + i, 8, k); c.fill = FILL_INPUT; c.border = BORDER; c.font = F_BASE
    c = R.cell(4 + i, 9, v); c.fill = FILL_INPUT; c.border = BORDER; c.font = F_BASE
R.cell(4 + len(NAMED_EXCL), 8, "Pakenham (HCMT - East Pakenham Depot) is not in the Site Register's large market rows; "
       "it already carries LGCS_HCMT and the SEC VIC 100% Renewable Deduction in Envizi.").font = F_NOTE
EXCL_RNG = (f"$H$4:$H${3+len(NAMED_EXCL)}", f"$I$4:$I${3+len(NAMED_EXCL)}")

HR = 18; FR = HR + 1; LR = HR + len(review)
REV_HEAD = ["#", "Register row", "Connection ID (NMI)", "State", "Green in register?", "Site Name (register)",
            "Address (register)", "Business Unit", "Contracted Retailer", "New Contract Start",
            "Envizi Account\n[formula]", "Envizi Location\n[formula]", "Location Ref\n[formula]",
            "Envizi Account Style\n[formula]", "Envizi Supplier\n[formula]",
            "Certificate accounts at location\n[formula]", "...of which on this NMI\n[formula]",
            "Certificate account names", "Green kWh Jul-Aug 26\n(account green component, renewal period) [formula]",
            "Consumption kWh Jun-Aug 26\n[formula]", "Suggested decision\n[formula]", "Decision\n[INPUT]",
            "Reason / notes", "New Account Number\n[formula]", "New Account Supplier\n[formula]",
            "Record Start\n[formula]", "Record End\n[formula]",
            "Virtual meter source account\n(the account carrying this NMI from 1 Jul 26) [formula]",
            "Source has data before 1 Jul 26?\n[formula]", "Opened On to set on the new account\n(contract start) [formula]"]
hdr(R, HR, REV_HEAD, height=58)
ACC_M, ACC_L, ACC_D, ACC_G, ACC_I = [f"{q(S_ACC)}!${c}$2:${c}${ACC_N}" for c in ("AL", "AL", "D", "G", "I")]
ACCJ = f"{q(S_ACC)}!$J$2:$J${ACC_N}"; ACCAJ = f"{q(S_ACC)}!$AJ$2:$AJ${ACC_N}"; ACCAM = f"{q(S_ACC)}!$AM$2:$AM${ACC_N}"; ACCAO = f"{q(S_ACC)}!$AO$2:$AO${ACC_N}"; ACCAP = f"{q(S_ACC)}!$AP$2:$AP${ACC_N}"
LOCB, LOCE = f"{q(S_LOC)}!$B$2:$B${LOC_N}", f"{q(S_LOC)}!$E$2:$E${LOC_N}"
SUMH, SUML, SUMW, SUMR = [f"{q(S_SUM)}!${c}$2:${c}${SUM_N}" for c in ("H", "L", "W", "R")]
rv = lambda col: f"{REV}!${col}${FR}:${col}${LR}"
SUP_FR = FR + len(review) + 3
R.cell(SUP_FR - 1, 33, "Contracted retailer -> Envizi supplier").font = F_NOTE
for i, (k, v) in enumerate(SUPPLIER_MAP.items()):
    R.cell(SUP_FR + i, 33, k).font = F_BASE
    R.cell(SUP_FR + i, 34, v).font = F_BASE
SUPMAP_K = f"{REV}!$AG${SUP_FR}:$AG${SUP_FR + len(SUPPLIER_MAP) - 1}"
SUPMAP_V = f"{REV}!$AH${SUP_FR}:$AH${SUP_FR + len(SUPPLIER_MAP) - 1}"

for j, x in enumerate(review):
    r = FR + j
    vals = [x["idx"], x["src_row"], x["nmi"], x["state"], x["green"], x["site"], x["addr"], x["bu"], x["contracted"], x["start"]]
    for c, v in enumerate(vals, start=1):
        cell = R.cell(r, c, v); cell.font = F_BASE; cell.border = BORDER; cell.alignment = TOP
    R.cell(r, 3).number_format = "@"; R.cell(r, 10).number_format = "yyyy-mm-dd"
    m = f"MATCH($C{r},{ACC_M},0)"
    f = {11: f'=IFERROR(INDEX({ACCJ},{m}),"Not found")',
         12: f'=IFERROR(INDEX({ACC_D},{m}),"Not found")',
         13: f'=IF(L{r}="Not found","",IFERROR(INDEX({LOCE},MATCH(L{r},{LOCB},0)),"Not in locations extract"))',
         14: f'=IFERROR(INDEX({ACC_G},{m}),"")',
         15: f'=IFERROR(INDEX({ACC_I},{m}),"")',
         16: f'=IF(L{r}="Not found",0,COUNTIF({ACCAM},L{r}))',
         17: f'=IF(L{r}="Not found",0,COUNTIFS({ACCAM},L{r},{ACCAJ},$C{r}))',
         19: f'=IF(K{r}="Not found",0,SUMIFS({SUMW},{SUML},K{r},{SUMH},"{GREEN_TYPE}",{SUMR},">="&DATE(2026,7,1)))',
         20: f'=IF(K{r}="Not found",0,SUMIFS({SUMW},{SUML},K{r},{SUMH},"{ELEC_TYPE}"))',
         21: (f'=IF(IFERROR(INDEX({EXCL_RNG[1]},MATCH($C{r},{EXCL_RNG[0]},0)),"")<>"","{D_NAMED}",'
              f'IF(AB{r}="None","{D_HOLD}",IF(AND(P{r}>0,{IN_LGCMODE}="{LGC_EMPTY}"),"{D_LGC}",'
              f'IF(S{r}>0,"{D_GREEN}",'
              f'IF(COUNTIF({ACCAP},$C{r}&"|"&IFERROR(INDEX({SUPMAP_V},MATCH(I{r},{SUPMAP_K},0)),I{r}))=0,"{D_TEMP}",'
              f'IF(OR(M{r}="Not in locations extract",M{r}=""),"{D_HOLD}","{D_CREATE}"))))))'),
         24: f'=IF(OR(V{r}="{D_CREATE}",V{r}="{D_TEMP}"),IF({IN_PATTERN}="{PAT_LGC}","LGCS_"&C{r},{IN_PREFIX}&AB{r}&{IN_SUFFIX}),"")',
         28: f'=IFERROR(INDEX({ACCJ},MATCH($C{r},{ACCAO},0)),"None")',
         29: f'=IF(AB{r}="Not found","",IF(SUMIFS({SUMW},{SUML},AB{r},{SUMH},"{ELEC_TYPE}",{SUMR},"<"&DATE(2026,7,1))>0,"Yes","No"))',
         25: f'=IF(OR(V{r}="{D_CREATE}",V{r}="{D_TEMP}"),IF({IN_PATTERN}="{PAT_LGC}","LGCs",{IN_NEWSUP}),"")',
         30: f'=IF(OR(V{r}="{D_CREATE}",V{r}="{D_TEMP}"),J{r},"")',
         26: f'=IF(OR(V{r}="{D_CREATE}",V{r}="{D_TEMP}"),J{r},"")',
         27: f'=IF(OR(V{r}="{D_CREATE}",V{r}="{D_TEMP}"),EOMONTH(J{r},0),"")'}
    for c, formula in f.items():
        cell = R.cell(r, c, formula); cell.font = F_BASE; cell.border = BORDER; cell.alignment = TOP
    R.cell(r, 18, x["cert_names"]).font = F_BASE
    c = R.cell(r, 22, x["decision"]); c.fill = FILL_INPUT; c.font = F_BASE
    c = R.cell(r, 23, x["reason"]); c.font = F_BASE; c.alignment = WRAP
    for c in (18, 22, 23):
        R.cell(r, c).border = BORDER; R.cell(r, c).alignment = WRAP
    for c in (19, 20):
        R.cell(r, c).number_format = "#,##0"
    for c in (26, 27):
        R.cell(r, c).number_format = "yyyy-mm-dd"
    for c in (28, 29, 30):
        R.cell(r, c).alignment = TOP
    R.cell(r, 30).number_format = "yyyy-mm-dd"
    R.row_dimensions[r].height = 45
dv_dec = DataValidation(type="list", formula1='"' + ",".join(DECISIONS) + '"', allow_blank=False,
                        error="Pick a decision from the list.", errorTitle="Decision")
R.add_data_validation(dv_dec); dv_dec.add(f"V{FR}:V{LR}")
rng = f"A{FR}:AD{LR}"
R.conditional_formatting.add(rng, FormulaRule(formula=[f'$V{FR}="{D_CREATE}"'], fill=FILL_GREEN, stopIfTrue=False))
R.conditional_formatting.add(rng, FormulaRule(formula=[f'$V{FR}="{D_TEMP}"'], fill=FILL_AMBER, stopIfTrue=False))
R.conditional_formatting.add(rng, FormulaRule(formula=[f'$V{FR}="{D_HOLD}"'], fill=FILL_AMBER, stopIfTrue=False))
R.conditional_formatting.add(rng, FormulaRule(formula=[f'LEFT($V{FR},7)="Exclude"'], fill=FILL_GREY, stopIfTrue=False))
R.conditional_formatting.add(f"V{FR}:V{LR}", FormulaRule(formula=[f'$V{FR}<>$U{FR}'], fill=FILL_RED,
                                                         font=Font(name=FONT, bold=True), stopIfTrue=False))
R.conditional_formatting.add(f"AC{FR}:AC{LR}", FormulaRule(formula=[f'$AC{FR}="Yes"'], fill=FILL_AMBER, stopIfTrue=False))
R.conditional_formatting.add(f"M{FR}:M{LR}", FormulaRule(formula=[f'OR($M{FR}="Not in locations extract",$M{FR}="")'],
                                                         fill=FILL_AMBER, stopIfTrue=False))
R.freeze_panes = f"D{FR}"
R.auto_filter.ref = f"A{HR}:AD{LR}"
widths(R, {"A": 5, "B": 8, "C": 16, "D": 7, "E": 9, "F": 30, "G": 34, "H": 11, "I": 12, "J": 12, "K": 28, "L": 34,
           "M": 16, "N": 22, "O": 14, "P": 12, "Q": 12, "R": 34, "S": 15, "T": 15, "U": 26, "V": 26, "W": 60,
           "X": 32, "Y": 18, "Z": 12, "AA": 12, "AB": 30, "AC": 14, "AD": 16})
R.sheet_properties.tabColor = "FF1F3864"

# ---------------------------------------------------------------- load + prep tabs
LOAD_HEAD = ["Organization Link", "Organization", "Location", "Location Ref", "Account Style Link",
             "Account Style Caption", "Account Subtype", "Account Number", "Account Reference", "Account Supplier",
             "Account Reader", "Record Start YYYY-MM-DD", "Record End YYYY-MM-DD", "Record Data Quality",
             "Record Billing Type", "Record Subtype", "Record Entry Method", "Record Reference",
             "Record Invoice Number", "Quantity", "Total Cost"]
for c, h in enumerate(LOAD_HEAD, start=1):
    ws_load.cell(1, c, h).font = F_BOLD
    ws_prep.cell(1, c, h).font = F_BOLD
ws_prep.cell(1, 22, "Key: Connection ID (NMI)").font = F_BOLD
REV = q(S_REV)
style_link_formula = f'=IF({REV}!{IN_STYLE}="","",{REV}!{IN_STYLE})'
for j, x in enumerate(creates):
    r = 2 + j
    row = [ORG_LINK, ORG, x["location"], x["lref"], None, STYLE_CAPTION, None, x["new_acct"], x["nmi"], x["new_sup"],
           None, x["rec_start"], x["rec_end"], "Actual", "Standard", None, "Overwrite", REC_REF, None, 0, None]
    for c, v in enumerate(row, start=1):
        cell = ws_load.cell(r, c, v); cell.font = F_BASE
    ws_load.cell(r, 5, style_link_formula).font = F_BASE
    for c in (8, 10, 11):   # account number, supplier, reader follow the pattern switch through Prep
        ws_load.cell(r, c, f"={q(S_PREP)}!{L(c)}{r}").font = F_BASE
    ws_load.cell(r, 9).number_format = "@"
    ws_load.cell(r, 12).number_format = "yyyy-mm-dd"; ws_load.cell(r, 13).number_format = "yyyy-mm-dd"
    ws_load.cell(r, 20).number_format = "0.0"
    # prep formulas keyed on col V
    k = f"MATCH($V{r},{rv('C')},0)"
    pf = {1: f"={REV}!{IN_ORG}", 2: f"={REV}!{IN_ORGNAME}",
          3: f'=IFERROR(INDEX({rv("L")},{k}),"")',
          4: f'=IFERROR(INDEX({LOCE},MATCH(C{r},{LOCB},0)),"")',
          5: style_link_formula, 6: f"={REV}!{IN_CAPTION}",
          8: f'=IFERROR(INDEX({rv("X")},{k}),"")', 9: f"=$V{r}",
          10: f'=IFERROR(INDEX({rv("Y")},{k}),"")',
          11: f'=IF({REV}!{IN_PATTERN}="{PAT_LGC}","LGCs",IF({REV}!{IN_READER}="","",{REV}!{IN_READER}))',
          12: f'=IFERROR(INDEX({rv("Z")},{k}),"")', 13: f'=IFERROR(INDEX({rv("AA")},{k}),"")',
          14: "Actual", 15: "Standard", 17: f"={REV}!{IN_METHOD}", 18: f"={REV}!{IN_REF}", 20: f"={REV}!{IN_QTY}",
          22: x["nmi"]}
    for c, v in pf.items():
        cell = ws_prep.cell(r, c, v); cell.font = F_BASE
    ws_prep.cell(r, 22).number_format = "@"
    ws_prep.cell(r, 12).number_format = "yyyy-mm-dd"; ws_prep.cell(r, 13).number_format = "yyyy-mm-dd"
    ws_prep.cell(r, 20).number_format = "0.0"
LOAD_LAST = 1 + len(creates)
for ws in (ws_load, ws_prep):
    widths(ws, {"A": 16, "B": 12, "C": 40, "D": 16, "E": 17, "F": 28, "G": 15, "H": 34, "I": 17, "J": 16, "K": 14,
                "L": 22, "M": 22, "N": 18, "O": 17, "P": 14, "Q": 19, "R": 42, "S": 20, "T": 10, "U": 10, "V": 22})
    ws.freeze_panes = "A2"
ws_load.sheet_properties.tabColor = "FF00B050"
ws_prep.cell(LOAD_LAST + 2, 1, "To add a row: copy a row down, put the Connection ID in col V, and the fields fill "
             "themselves from the Review tab. Paste the row as values onto the load tab.").font = F_NOTE

# ---------------------------------------------------------------- Check tab
C = ws_chk
C["A1"] = "Checks - all should read OK before the load tab goes to Envizi"; C["A1"].font = F_TITLE
hdr(C, 3, ["Check", "Value", "Expected", "Status"], height=22)
LOADH = f"{q(S_LOAD)}!$H$2:$H${LOAD_LAST}"; LOADD = f"{q(S_LOAD)}!$D$2:$D${LOAD_LAST}"
LOADC = f"{q(S_LOAD)}!$C$2:$C${LOAD_LAST}"; LOADL = f"{q(S_LOAD)}!$L$2:$L${LOAD_LAST}"; LOADM = f"{q(S_LOAD)}!$M$2:$M${LOAD_LAST}"
PREPH = f"{q(S_PREP)}!$H$2:$H${LOAD_LAST}"; PREPD = f"{q(S_PREP)}!$D$2:$D${LOAD_LAST}"; PREPC = f"{q(S_PREP)}!$C$2:$C${LOAD_LAST}"
PREPL = f"{q(S_PREP)}!$L$2:$L${LOAD_LAST}"; PREPK = f"{q(S_PREP)}!$K$2:$K${LOAD_LAST}"; LOADK = f"{q(S_LOAD)}!$K$2:$K${LOAD_LAST}"; PREPM = f"{q(S_PREP)}!$M$2:$M${LOAD_LAST}"; PREPJ = f"{q(S_PREP)}!$J$2:$J${LOAD_LAST}"; LOADJ = f"{q(S_LOAD)}!$J$2:$J${LOAD_LAST}"
RV = f"{REV}!$V${FR}:$V${LR}"; RU = f"{REV}!$U${FR}:$U${LR}"; RC = f"{REV}!$C${FR}:$C${LR}"
checks = [
    ("Site Register rows in scope (AU Large Electricity)", f"=COUNTA({RC})", len(register)),
    (f"Decision = {D_CREATE}", f'=COUNTIF({RV},"{D_CREATE}")', None),
    (f"Decision = {D_TEMP}", f'=COUNTIF({RV},"{D_TEMP}")', None),
    (f"Decision = {D_LGC}", f'=COUNTIF({RV},"{D_LGC}")', None),
    (f"Decision = {D_GREEN}", f'=COUNTIF({RV},"{D_GREEN}")', None),
    (f"Decision = {D_NAMED}", f'=COUNTIF({RV},"{D_NAMED}")', None),
    (f"Decision = {D_HOLD}", f'=COUNTIF({RV},"{D_HOLD}")', None),
    ("Decisions add up to the rows in scope", "=SUM(B5:B10)", "=B4"),
    ("Rows where Decision differs from Suggested (review these)", f"=SUMPRODUCT(({RV}<>{RU})*1)", 0),
    ("Rows on the load tab", f"=COUNTA({LOADH})", "=B5+B6"),
    ("Duplicate account numbers on the load tab", f"=SUMPRODUCT((COUNTIF({LOADH},{LOADH})>1)*1)", 0),
    ("Load rows with no Location Ref", f'=COUNTIF({LOADD},"")', 0),
    ("Load rows already built in Envizi (account exists in the accounts extract)", f"=SUMPRODUCT(COUNTIF({ACCJ},{LOADH}))", None),
    ("Account Style Link filled on the Review tab (Envizi rejects the file without it)", f'=IF({REV}!{IN_STYLE}="","NOT FILLED","Filled")', "Filled"),
    ("Prep vs load - Location", f"=SUMPRODUCT(({PREPC}<>{LOADC})*1)", 0),
    ("Prep vs load - Location Ref", f"=SUMPRODUCT(({PREPD}<>{LOADD})*1)", 0),
    ("Prep vs load - Account Number", f"=SUMPRODUCT(({PREPH}<>{LOADH})*1)", 0),
    ("Prep vs load - Account Supplier", f"=SUMPRODUCT(({PREPJ}<>{LOADJ})*1)", 0),
    ("Prep vs load - Account Reader", f"=SUMPRODUCT(({PREPK}&\"\"<>{LOADK}&\"\")*1)", 0),
    ("Prep vs load - Record Start", f"=SUMPRODUCT(({PREPL}<>{LOADL})*1)", 0),
    ("Prep vs load - Record End", f"=SUMPRODUCT(({PREPM}<>{LOADM})*1)", 0),
]
for i, (label, formula, expected) in enumerate(checks):
    r = 4 + i
    C.cell(r, 1, label).font = F_BASE
    C.cell(r, 2, formula).font = F_BASE
    if expected is not None:
        C.cell(r, 3, expected).font = F_BASE
        C.cell(r, 4, f'=IF(B{r}=C{r},"OK","CHECK")').font = F_BOLD
    else:
        C.cell(r, 4, "info").font = F_NOTE
    for c in range(1, 5):
        C.cell(r, c).border = BORDER
CH_LAST = 3 + len(checks)
C.conditional_formatting.add(f"D4:D{CH_LAST}", FormulaRule(formula=['$D4="OK"'], fill=FILL_GREEN))
C.conditional_formatting.add(f"D4:D{CH_LAST}", FormulaRule(formula=['$D4="CHECK"'], fill=FILL_RED))
widths(C, {"A": 78, "B": 16, "C": 12, "D": 10})


# ---------------------------------------------------------------- Manual Setup Checklist
M = ws_man
M["A1"] = "Setting the virtual accounts up by hand - one per qualifying large market account"; M["A1"].font = F_TITLE
M.row_dimensions[1].height = 22
STEPS = [
 "A virtual meter has to be empty. Envizi only lets an account be set up as a virtual meter - its records calculated from a source account - "
 "while it holds no records, and the bulk PM&C template cannot create an account without a record. So the load tab is not uploaded for this; "
 "each account is created and linked in Envizi by hand, using the rows below as the worklist.",
 "STEP 1  Open the location (Manage > Locations, or from the source account in col E) and Add Account. Account style: Certificates - Location - kWh. "
 "Account number, reference, supplier and reader as shown in cols H-L, and Opened On = the date in col S (2026-07-01) - that is what stops the meter "
 "reaching back before the renewal. Save without adding any record. Only the register's green-row NMIs get a virtual meter; any other electricity "
 "account at the same location is left alone.",
 "STEP 2  Open the new account and set it up as a virtual meter. Source account = the electricity account in col E, 100%, add. Link that one "
 "only. Where col M lists another account on the same NMI it is a leftover that was never closed, not a second live supply - on 13 of these "
 "NMIs it is still recording the months the source already covers, which is a double count to close off on the electricity side.",
 "STEP 3  Check that the new account now shows kWh equal to the source for the latest months and that the location's market-based CO2e has "
 "dropped to match. Mark Status (col N) and Date (col O). Anything odd goes in col P.",
 "STEP 4  Where col R names an existing LGCS_ account, leave it alone - it holds 2025-and-earlier LGC data (checked 4 Sep 26), so it cannot "
 "become a virtual meter and stays as the historical record. The new account sits beside it; that is also why the new names are keyed on the "
 "source account rather than LGCS_<NMI>.",
 "HISTORY  A virtual meter mirrors every period its source has data for. Where col Q says Yes the source account has data before 1 Jul 26; the "
 "Opened On of 2026-07-01 on the new account bounds it - the first two (Bathurst, Mogo 4001127731) show July and August only, with June data on "
 "the source untouched. Verified against the 4 Sep 26 export: kWh identical to the source, offset booked as negative Other CO2e. One thing to fix "
 "first: the certificate rows use the 'LGCs NSW 24-25' factor (-0.66) against electricity on '25-26' (0.64), so NSW sites net about 3% negative - "
 "get a 25-26 LGC factor mapped for NSW before the rest are linked.",
 "Existing LGCS_ accounts carry Usage Type 'Consolidation', Use 'CO2e and Base Measure' and Apportionment 0 - match those if the screen offers them, "
 "otherwise leave the defaults. The Account pattern input on the Review tab (C14) drives the names below.",
]
for i, t in enumerate(STEPS):
    c = M.cell(3 + i, 1, t); c.font = F_BASE if i else F_NOTE; c.alignment = WRAP
    M.merge_cells(start_row=3 + i, start_column=1, end_row=3 + i, end_column=19)
    M.row_dimensions[3 + i].height = 46 if i else 44
MH = 3 + len(STEPS) + 1; MF = MH + 1
MAN_HEAD = ["#", "Connection ID (NMI)", "Location", "Location Ref", "Source account to link\n(virtual meter source)",
            "Source account supplier", "Contracted retailer", "New account number", "Account Style", "Account Reference",
            "Account Supplier", "Account Reader", "Other accounts on this NMI\n(NOT linked - close off)", "Status\n[INPUT]",
            "Date actioned\n[INPUT]", "Notes\n[INPUT]", "Source has data before 1 Jul 26?\n(date-bound the meter)",
            "Historical LGC account(s) at this location\n(keep - not reusable)", "Opened On to set\n(bounds the meter to the renewal)"]
hdr(M, MH, MAN_HEAD, height=58)
# other electricity accounts on the same NMI (incoming retailer, possibly still unallocated)
def other_accounts(nmi, chosen):
    rows = keep[(keep["_kind"] == 0) & (keep["_nmi"] == nmi) & (keep["_rep"] == "") & (keep["Account Number"] != chosen)]
    out = []
    for _, a in rows.iterrows():
        tag = " (still at Unallocated Accounts)" if a["Location"] == "Unallocated Accounts" else ""
        out.append(f"{a['Account Number']} - {a['Supplier']}{tag}")
    return out
for j, x in enumerate(creates):
    r = MF + j
    k = f"MATCH($B{r},{rv('C')},0)"
    others = [x["other_on_nmi"]] if x["other_on_nmi"] else []
    vals = {1: j + 1, 2: x["nmi"],
            3: f'=IFERROR(INDEX({rv("L")},{k}),"")', 4: f'=IFERROR(INDEX({rv("M")},{k}),"")',
            5: f'=IFERROR(INDEX({rv("AB")},{k}),"")', 6: f'{x["src_sup"]} · {x["src_style"]}',
            7: f'=IFERROR(INDEX({rv("I")},{k}),"")', 8: f'=IFERROR(INDEX({rv("X")},{k}),"")',
            9: f"={REV}!{IN_CAPTION}", 10: f"=$B{r}", 11: f'=IFERROR(INDEX({rv("Y")},{k}),"")',
            12: f'=IF({REV}!{IN_PATTERN}="{PAT_LGC}","LGCs",IF({REV}!{IN_READER}="","",{REV}!{IN_READER}))',
            13: "; ".join(others), 14: "Not started", 15: None, 16: None,
            17: f'=IFERROR(INDEX({rv("AC")},{k}),"")', 18: x["cert_names"], 19: f'=IFERROR(INDEX({rv("AD")},{k}),"")'}
    for c, v in vals.items():
        cell = M.cell(r, c, v); cell.font = F_BASE; cell.border = BORDER; cell.alignment = WRAP if c in (3, 5, 13, 16, 18) else TOP
    M.cell(r, 2).number_format = "@"; M.cell(r, 15).number_format = "dd-mmm-yyyy"; M.cell(r, 19).number_format = "yyyy-mm-dd"
    for c in (14, 15, 16): M.cell(r, c).fill = FILL_INPUT
    M.row_dimensions[r].height = 32
ML = MF + len(creates) - 1
STATUS_LIST = '"Not started,Account created,Linked as virtual meter,Verified,Not proceeding"'
dv_st = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=True); M.add_data_validation(dv_st); dv_st.add(f"N{MF}:N{ML}")
rngm = f"A{MF}:S{ML}"
M.conditional_formatting.add(rngm, FormulaRule(formula=[f'$N{MF}="Verified"'], fill=FILL_GREEN))
M.conditional_formatting.add(rngm, FormulaRule(formula=[f'$N{MF}="Linked as virtual meter"'], fill=FILL_SUB))
M.conditional_formatting.add(rngm, FormulaRule(formula=[f'$N{MF}="Account created"'], fill=FILL_AMBER))
M.conditional_formatting.add(rngm, FormulaRule(formula=[f'$N{MF}="Not proceeding"'], fill=FILL_GREY))
M.conditional_formatting.add(f"Q{MF}:Q{ML}", FormulaRule(formula=[f'$Q{MF}="Yes"'], fill=FILL_AMBER))
M.freeze_panes = f"C{MF}"; M.auto_filter.ref = f"A{MH}:S{ML}"
widths(M, {"A": 5, "B": 16, "C": 34, "D": 14, "E": 30, "F": 14, "G": 12, "H": 32, "I": 26, "J": 14, "K": 14, "L": 10,
           "M": 46, "N": 22, "O": 14, "P": 40, "Q": 16, "R": 40, "S": 16})
M.sheet_properties.tabColor = "FF00B050"

# ---------------------------------------------------------------- LGCS Accounts to Check
G = ws_lgc
G["A1"] = "Existing certificate accounts at the in-scope sites - all hold data, so none can be reused as a virtual meter"; G["A1"].font = F_TITLE
G["A2"] = ("Every active Certificates - Location [kWh] account at an in-scope location. Checked in Envizi on 4 Sep 26: all of them hold LGC data from "
           "2025 and earlier, so none is empty and none can be converted to a virtual meter - they stay as the historical record and a new virtual "
           "account is created beside each (Manual Setup Checklist). Col I is pre-set to Yes; if one turns out to be empty after all, set it to No and "
           "the Action column switches to reuse (open the account, set it up as a virtual meter with the live electricity account(s) in col G as "
           "source(s) at 100%).")
G["A2"].font = F_NOTE; G["A2"].alignment = WRAP; G.merge_cells("A2:L2"); G.row_dimensions[2].height = 58
GH = 4; GF = GH + 1
LGC_HEAD = ["#", "Certificate account", "Location", "Location Ref", "Supplier / Reader", "Register NMI(s) at this location",
            "Live electricity account(s) at the location to link", "On the account's own NMI?", "Has data?\n[INPUT]",
            "Action\n[formula]", "Status\n[INPUT]", "Notes\n[INPUT]"]
hdr(G, GH, LGC_HEAD, height=48)
excl_locs = {}
for x in review:
    if x["n_cert_loc"] > 0 and x["decision"] in MAKE:
        excl_locs.setdefault(x["location"], {"lref": x["lref"], "nmis": []})["nmis"].append(x["nmi"])
lgc_rows = []
for L_, info in excl_locs.items():
    c_here = keep[(keep["_kind"] == 1) & (keep["_active"] == 1) & (keep["Location"] == L_) & (keep["Supplier"] != NEW_SUPPLIER)]
    live = keep[(keep["_kind"] == 0) & (keep["_rep"] == "") & (keep["_nmi"].isin(info["nmis"]))]
    live_txt = "; ".join(f"{a['Account Number']} - {a['Supplier']}" + (" (still at Unallocated Accounts)" if a["Location"] == "Unallocated Accounts" else "")
                         for _, a in live.sort_values("Account Number").iterrows())
    for _, c in c_here.sort_values("Account Number").iterrows():
        lgc_rows.append(dict(acct=c["Account Number"], loc=L_, lref=info["lref"], sup=f"{c['Supplier']} / {c['Reader']}".strip(" /"),
                             nmis=", ".join(info["nmis"]), live=live_txt, own=("Y" if c["_nmi"] in info["nmis"] else "N")))
lgc_rows.sort(key=lambda d: (d["loc"], d["acct"]))
for j, d in enumerate(lgc_rows):
    r = GF + j
    vals = {1: j + 1, 2: d["acct"], 3: d["loc"], 4: d["lref"], 5: d["sup"], 6: d["nmis"], 7: d["live"], 8: d["own"], 9: "Yes",
            10: (f'=IF(I{r}="No","Reuse - set up as the virtual meter of the account(s) in col G",'
                 f'IF(I{r}="Yes","Has data (2025 and earlier) - keep as the historical record; the new virtual account for this site is on the Manual Setup Checklist",'
                 f'"Check in Envizi - Records / First and Last Record"))'),
            11: "Not started", 12: None}
    for c, v in vals.items():
        cell = G.cell(r, c, v); cell.font = F_BASE; cell.border = BORDER; cell.alignment = WRAP if c in (3, 6, 7, 10, 12) else TOP
    for c in (9, 11, 12): G.cell(r, c).fill = FILL_INPUT
    G.row_dimensions[r].height = 32
GL = GF + len(lgc_rows) - 1
dv_hd = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True); G.add_data_validation(dv_hd); dv_hd.add(f"I{GF}:I{GL}")
dv_gs = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=True); G.add_data_validation(dv_gs); dv_gs.add(f"K{GF}:K{GL}")
rngg = f"A{GF}:L{GL}"
G.conditional_formatting.add(rngg, FormulaRule(formula=[f'$K{GF}="Verified"'], fill=FILL_GREEN))
G.conditional_formatting.add(rngg, FormulaRule(formula=[f'$K{GF}="Linked as virtual meter"'], fill=FILL_SUB))
G.conditional_formatting.add(f"I{GF}:J{GL}", FormulaRule(formula=[f'$I{GF}="Yes"'], fill=FILL_AMBER))
G.conditional_formatting.add(f"I{GF}:J{GL}", FormulaRule(formula=[f'$I{GF}="No"'], fill=FILL_GREEN))
G.conditional_formatting.add(f"H{GF}:H{GL}", FormulaRule(formula=[f'$H{GF}="N"'], fill=FILL_AMBER))
G.freeze_panes = f"C{GF}"; G.auto_filter.ref = f"A{GH}:L{GL}"
widths(G, {"A": 5, "B": 30, "C": 34, "D": 14, "E": 14, "F": 24, "G": 50, "H": 12, "I": 11, "J": 46, "K": 22, "L": 36})
G.sheet_properties.tabColor = "FF1F3864"
print("checklist rows:", len(creates), "| LGCS accounts to check:", len(lgc_rows), "at", len(excl_locs), "locations")

# ---------------------------------------------------------------- Site Register tab
hdr(ws_reg, 1, REG_HEADERS + ["Green highlighted", "Source row"], height=30)
for j, x in enumerate(register):
    r = 2 + j
    for c, v in enumerate(x["vals"], start=1):
        cell = ws_reg.cell(r, c, v); cell.font = F_BASE
        if isinstance(v, datetime.datetime): cell.number_format = "yyyy-mm-dd"
    ws_reg.cell(r, 1).number_format = "@"
    ws_reg.cell(r, 22, "Y" if x["green"] else "N").font = F_BASE
    ws_reg.cell(r, 23, x["src_row"]).font = F_BASE
    if x["green"]:
        for c in range(1, 22): ws_reg.cell(r, c).fill = PatternFill("solid", start_color="FF92D050", end_color="FF92D050")
ws_reg.freeze_panes = "B2"; ws_reg.auto_filter.ref = f"A1:W{1+len(register)}"
widths(ws_reg, {"A": 16, "B": 12, "C": 11, "D": 8, "E": 20, "F": 32, "G": 44, "H": 40, "I": 12, "J": 7, "K": 14,
                "L": 14, "M": 20, "N": 12, "O": 12, "P": 11, "Q": 12, "R": 12, "S": 12, "T": 15, "U": 24, "V": 10, "W": 9})
ws_reg.cell(len(register) + 3, 1, "Values copied from 'Site Register' in Downer_Energy_Contracting_and_Budget_Summary_FY26-28.xlsx, "
            "AU Large Electricity rows only. Green fill reproduced from the source.").font = F_NOTE

# ---------------------------------------------------------------- Accounts extract tab (filtered + helpers)
acc_head = ACC_HEADERS + ["_NMI\n(text after the last underscore)", "_Active\n(1 = not replaced, not Unallocated)",
                          "_Match\n(NMI if active electricity account)", "_CertLocation\n(location if active certificate account)",
                          "_Match_Supplier\n(NMI|Supplier for active electricity accounts)",
                          "_Match_LM\n(NMI if active LARGE MARKET electricity account)"]
hdr(ws_acc, 1, acc_head, height=44)
DATE_IDX = {14, 15, 19, 20, 24, 25, 30, 31, 35}
INT_IDX = {1, 2}; NUM_IDX = {18, 23, 29}
for j, (_, rec) in enumerate(keep.iterrows()):
    r = 2 + j
    for c, name in enumerate(ACC_HEADERS, start=1):
        raw = rec[name]
        if c in DATE_IDX:
            val = parse_date(raw)
        elif c in INT_IDX or c in NUM_IDX:
            try: val = float(raw); val = int(val) if val.is_integer() else val
            except ValueError: val = raw or None
        else:
            val = raw.strip() or None
        cell = ws_acc.cell(r, c, val); cell.font = F_BASE
        if c in DATE_IDX: cell.number_format = "dd-mmm-yyyy"
    ws_acc.cell(r, 36, f'=IFERROR(MID(J{r},FIND("~",SUBSTITUTE(J{r},"_","~",LEN(J{r})-LEN(SUBSTITUTE(J{r},"_",""))))+1,LEN(J{r})),J{r})').font = F_BASE
    ws_acc.cell(r, 37, f'=IF(AND(OR(ISBLANK(O{r}),O{r}=""),D{r}<>"Unallocated Accounts"),1,0)').font = F_BASE
    ws_acc.cell(r, 38, f'=IF(AND(AK{r}=1,F{r}="{ELEC_TYPE}"),AJ{r},"")').font = F_BASE
    ws_acc.cell(r, 39, f'=IF(AND(AK{r}=1,F{r}="{CERT_TYPE}"),D{r},"")').font = F_BASE
    ws_acc.cell(r, 40, f'=IF(AL{r}<>"",AL{r}&"|"&I{r},"")').font = F_BASE
    ws_acc.cell(r, 41, f'=IF(AL{r}<>"",AJ{r},"")').font = F_BASE
    ws_acc.cell(r, 42, f'=IF(AL{r}<>"",AJ{r}&"|"&I{r},"")').font = F_BASE
ACC_LAST = 1 + len(keep)
ws_acc.freeze_panes = "A2"; ws_acc.auto_filter.ref = f"A1:AO{ACC_LAST}"
widths(ws_acc, {"D": 36, "E": 16, "F": 26, "G": 28, "I": 14, "J": 34, "K": 16, "N": 12, "O": 12, "AI": 12,
                "AJ": 18, "AK": 14, "AM": 30, "AN": 26, "AO": 26})
ws_acc.cell(ACC_LAST + 2, 1, f"Rows from Extract_for_Accounts 05 Sep 26.csv: every Electricity [kWh] account on one of the "
            f"{len(register)} register NMIs (any status), plus every {CERT_TYPE} account. Sorted so the preferred match "
            "(Large Market style, active) comes first for each NMI - the Review lookups take the first match. "
            "An empty Replaced On is exported as 30 Dec 1899 and is loaded here as blank.").font = F_NOTE

# ---------------------------------------------------------------- Locations extract tab (full)
hdr(ws_loc, 1, LOC_HEADERS, height=30)
for j, rec in enumerate(loc_body):
    r = 2 + j
    for c, raw in enumerate(rec, start=1):
        v = raw.strip() if isinstance(raw, str) else raw
        if v == "": continue
        ws_loc.cell(r, c, v)
ws_loc.freeze_panes = "A2"
widths(ws_loc, {"B": 44, "D": 18, "E": 18, "H": 30, "I": 16, "N": 14, "R": 30})

# ---------------------------------------------------------------- Summary tab (filtered)
SUM_HEAD = list(summ_keep.columns)
hdr(ws_sum, 1, SUM_HEAD, height=30)
for j, (_, rec) in enumerate(summ_keep.iterrows()):
    r = 2 + j
    for c, name in enumerate(SUM_HEAD, start=1):
        v = rec[name]
        if v == "" or v is None: continue
        if isinstance(v, pd.Timestamp): v = v.to_pydatetime()
        cell = ws_sum.cell(r, c, v)
        if isinstance(v, datetime.datetime): cell.number_format = "yyyy-mm-dd"
ws_sum.freeze_panes = "A2"; ws_sum.auto_filter.ref = f"A1:{L(len(SUM_HEAD))}{1+len(summ_keep)}"
widths(ws_sum, {"E": 36, "H": 26, "J": 30, "L": 26, "N": 12, "R": 12, "W": 14, "AI": 12})
ws_sum.cell(len(summ_keep) + 3, 1, "Rows from ElectricityEnviziSummaryjunejulyaug26.xlsx whose Item Number sits on one of the "
            "register NMIs - both the Electricity [kWh] and the Electricity - Green [kWh] components.").font = F_NOTE

# ---------------------------------------------------------------- Notes tab
N = ws_not
notes = [
 ("WHAT THIS FILE IS", True),
 ("The Envizi Account Setup and Data Load (PM&C) upload that creates one virtual certificate account at each AU large market "
  "location that does not already have an offset. Upload the first tab only. The filename keeps the "
  "Account_Setup_and_Data_Load_-_PM&C_ prefix Envizi processes on.", False),
 ("", False),
 ("SOURCES", True),
 ("Site Register - Downer_Energy_Contracting_and_Budget_Summary_FY26-28.xlsx (81 AU Large Electricity rows; 78 green, the 3 NT rows not).", False),
 ("Accounts extract - FY27/Extract_for_Accounts 05 Sep 26.csv.  Locations extract - FY27/Extract_for_Locations 26 Aug 26.csv.", False),
 ("Monthly summary - ElectricityEnviziSummaryjunejulyaug26.xlsx (Jun-Aug 26 electricity and green components, with cost).", False),
 ("", False),
 ("HOW EACH SITE WAS MATCHED", True),
 ("On the connection ID. The NMI is the text after the last underscore in the Envizi account number, so each register row is matched to "
  "the active account carrying that NMI's electricity from 1 Jul 26 - actual data first, Large Market style as the tie-break - and the location "
  "comes from that account. Region was not used for matching. The monthly summary confirms each matched account is the one billing "
  "at that location. All 81 rows resolved to an account; the five NMIs that also have a new unallocated account (Sep-26 tracker) take the allocated one.", False),
 ("", False),
 ("SCOPE AND EXCLUSIONS", True),
 ("SCOPE FOLLOWS THE SUPPLY AGREEMENT, NOT THE METER CLASS. Category Management, 4 Sep 26, on a Mogo NMI: 'sometimes when you take both to a "
  "retailer they will agree to supply both as large market sites... from a metering perspective it may be considered a small site, but from an "
  "electricity supply agreement perspective, it's being treated as a large site.' So Electricity Small Market in Envizi is the metering "
  "classification and does not take a site out of the renewal - the register's green rows decide. 35 of the 70 accounts sit on a "
  "small-market-styled source for that reason.", False),
 ("Also in scope: the green AU Large Electricity rows. The three Northern Territory sites are not green and are named exclusions anyway; "
  "QTMP (Torbanlea, 3053253239) is a named exclusion. Pakenham (HCMT - East Pakenham Depot) is not in the register's large market rows "
  "and already carries LGCS_HCMT plus the SEC VIC 100% Renewable Deduction in Envizi.", False),
 ("Existing LGCS_<NMI> accounts: 52 rows at 40 locations already have one. Checked in Envizi on 4 Sep 26, all of them hold LGC data from 2025 "
  "and earlier, so none can become a virtual meter and none covers the renewal period - those sites get a new virtual account too (70 in all) and "
  "the old accounts stay as the historical record. The input at C15 switches back to excluding them if any turn out to be empty.", False),
 ("Excluded for an offset already in place: the seven WA Alinta accounts, whose own green component has been recording 100% green kWh since the "
  "1 Jul 26 contract - flip those to Create if a certificate account is wanted as well.", False),
 ("New Zealand: the 158 NZ green rows are all contracted to Ecotricity, but only 26 of those ICPs exist as electricity accounts in Envizi "
  "and 20 of them already have the Ecotricity _CERTS copy. They are not in this load; the six matched-but-uncovered Meridian sites are listed in the README.", False),
 ("", False),
 ("WHAT THE RATE SCHEDULE SAYS ABOUT LGCs", True),
 ("Rates & Source Data in the budget workbook prices the Engie and Origin renewal contracts with an explicit LGC line - FY27 0.395 c/kWh "
  "(Engie) and 0.375 (Origin), FY28 0.375 - plus a 'renewable product $/yr': 68 of the 78 green sites. The Alinta (WA) and Shell (TAS) "
  "contracts are modelled as an all-in delivered rate with no separate LGC line; Category Management confirms they are renewable, and the "
  "Alinta accounts already show 100% green kWh in Envizi. Small market sites are not renewable and are not in the rate schedule.", False),
 ("The existing LGCS_<NMI> accounts are the same mechanism recorded the Envizi way - 55 of the 69 were created in one batch on 12 Mar 2024 "
  "and 11 on 27 Jun 2025. Whether any of them hold data is not visible in the accounts extract or the electricity summary; a "
  "Certificates - Location data export would show it. An empty one is the virtual meter for that site, which is why locations that have "
  "one are excluded here rather than given a second account.", False),
 ("", False),
 ("A VIRTUAL METER HAS TO BE EMPTY", True),
 ("Envizi only lets an account be set up as a virtual meter while it holds no records, and this template cannot create an account without a "
  "record - the placeholder row on the load tab would leave each new account non-empty. Where the accounts are to be linked as virtual meters, the "
  "load tab is not uploaded: the Manual Setup Checklist tab is the worklist for creating and linking each one in Envizi by hand, and the "
  "LGCS Accounts to Check tab decides where an existing empty LGCS_ account is reused instead. The load tab stays as the record of what each "
  "account should look like, and can be uploaded if a setup-only route to empty accounts is confirmed.", False),
 ("", False),
 ("HOW THE NEW ACCOUNTS ARE STRUCTURED", True),
 ("Modelled on the Ecotricity virtual accounts (Copy of Eco_ICP_..._CERTS): same location as the large market account, account style "
  "Certificates - Location - kWh, account number = the virtual meter's source account + _CERTS (prefix and suffix are inputs on the Review tab; "
  "the 'Copy of ' that Envizi's copy function adds is left off), Account Reference = the NMI, Supplier = 'LGC Virtual Account' (input C16, as used "
  "on the first two), Reader blank, Opened On = the contract start (2026-07-01), which is what bounds the virtual meter to the renewal period. "
  "The source account is the contracted retailer's account on the NMI where one exists, otherwise the matched account. Only the register's "
  "green-row NMIs get a virtual meter; other electricity accounts at the same location are not under the agreement.", False),
 ("A virtual meter mirrors every period its source has data for. Review col AC flags the sources that have data before 1 Jul 26 (the continuing "
  "Origin accounts, and CS Energy where the Engie account is not yet in Envizi): date-bound those meters to the renewal start, or wait for the "
  "new-contract account, so no certificates are generated for periods that were not renewable.", False),
 ("The Account pattern input on the Review tab switches the whole set to the existing AU convention - LGCS_<NMI>, supplier and reader "
  "LGCs - in one step. Cols E, H, J and K of the load tab are formulas reading the Review and Prep tabs, so the style link and the "
  "pattern flow straight through; every other load column is a value.", False),
 ("Envizi will not create an account from this template without a record, so each row carries one placeholder record: the contract's first "
  "month (2026-07-01 to 2026-07-31), Quantity 0, Entry Method Overwrite - the real certificate quantity for that month replaces it when loaded.", False),
 ("Account Style Link is mandatory and is the one value not in any extract - enter the ID for Certificates - Location - kWh in the yellow "
  "cell on the Review tab (Envizi > Admin > Account Styles) and it flows to col E of the load and prep tabs. The Check tab reads NOT FILLED until then.", False),
 ("", False),
 ("BEFORE LOADING", True),
 ("1. Fill the Account Style Link.  2. Confirm the Decision column on the Review tab - anything changed from Suggested shows red.  "
  "3. Check tab all OK.  4. Recalculate (F9), save, upload the first tab.", False),
]
for i, (text, bold) in enumerate(notes):
    c = N.cell(1 + i, 1, text)
    c.font = F_BOLD if bold else F_BASE
    c.alignment = WRAP
    if not bold and text:
        N.row_dimensions[1 + i].height = 15 * (1 + len(text) // 150)
N.column_dimensions["A"].width = 150

# ---------------------------------------------------------------- properties, save
wb.properties.creator = ""
wb.properties.lastModifiedBy = ""
wb.properties.title = "Account Setup and Data Load - PM&C - Large market certificate accounts Jul-26"
wb.save(OUT)
print("saved", OUT)

# stash the python-side view for the README / summary
pd.DataFrame(review).to_csv("/tmp/claude-0/-home-user-RevenueAndUnallocated/3dd2248b-bb2f-5ae5-bbe4-76df52397984/scratchpad/review.csv", index=False)
